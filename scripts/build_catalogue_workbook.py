"""Build the empty Settlements catalogue workbook — one sheet per catalogue.

    py -3.13 scripts/build_catalogue_workbook.py

A clean input surface for rebuilding the catalogue layer by hand. Every sheet is
empty apart from a header row and ONE greyed example row showing the expected
format; delete the example when you start typing.

Two deliberate choices carried over from the audit:

  * Every priced catalogue has a **Conf** column (A/B/C confidence tier).
    Deployables and Structures never had one, which is how 15 unpriced
    deployables and 23 untiered structure costs shipped without anyone seeing
    it. "No number ships untagged" is the standard; the column enforces it.

  * Chems and Drones get sheets even though they do not exist in the rules
    today. They are named in the setting and referenced by built structures,
    so leaving them out would repeat the omission rather than fix it.
"""
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.normpath(os.path.join(HERE, "..", "catalogues",
                                    "SETTLEMENTS-CATALOGUES.xlsx"))

INK = "1F2933"
BAND = {"crew": "4A5568", "gear": "8A5A1B", "deploy": "7A4E1E",
        "combat": "8C2F26", "terrain": "2C5E46", "campaign": "6B5D1E",
        "meta": "2E5C7A"}

# (sheet, band, [(header, width), ...], example row)
SHEETS = [
    # ---- CREW ----------------------------------------------------------
    ("Ranks", "crew",
     [("Rank", 16), ("Stat Points", 12), ("Orders", 9), ("Credits", 10),
      ("Conf", 7), ("Tier Caps at Creation", 34), ("Notes", 46)],
     ["Fighter", 5, 0, 100, "A", "2 x T1", "The workhorse; first rank that may carry Standard Ranged"]),

    ("Skills", "crew",
     [("Skill", 24), ("Stat", 8), ("Tier", 7), ("Effect", 74), ("Role", 22), ("Notes", 30)],
     ["Bank Shot", "DEX", "T1", "When Heavy Cover causes your ranged attack to miss by exactly 1, a target touching that cover becomes Pinned; make no Injury roll.", "Suppressor", ""]),

    ("Progression", "crew",
     [("Level", 8), ("Grants", 44), ("Notes", 50)],
     [3, "Skill slot — any path that stat has unlocked", "First skill slot of the track"]),

    # ---- GEAR ----------------------------------------------------------
    ("Weapon Classes", "gear",
     [("Class", 20), ("Dmg Min", 9), ("Dmg Max", 9), ("Range Min", 10), ("Range Max", 10),
      ("Hands", 7), ("Min Rank", 12), ("Slots", 7), ("Base Cr", 9), ("Conf", 7),
      ("Always Has", 34)],
     ["Standard Ranged", 2, 4, 12, 36, 2, "Fighter", 3, 15, "A", "Two-Handed, Loud"]),

    ("Weapon Characteristics", "gear",
     [("Characteristic", 20), ("Category", 16), ("Cr", 8), ("Conf", 7),
      ("Restriction", 26), ("Effect", 74)],
     ["Brutal", "Damage & armour", 10, "A", "Melee, or ranged with Short Range",
      "+1 Damage, to the class band's ceiling (max +5)."]),

    ("Weapon Drawbacks", "gear",
     [("Drawback", 18), ("Refund", 9), ("Conf", 7), ("Restriction", 24), ("Effect", 62)],
     ["Slow", -5, "C", "Melee only", "You may not Charge with this weapon."]),

    ("Armour", "gear",
     [("Armour", 18), ("Injury Mod", 11), ("Cr", 8), ("Conf", 7), ("Drawback", 26), ("Notes", 44)],
     ["Light", -1, 10, "B", "None", "Measured directly with zero prior"]),

    ("Equipment", "gear",
     [("Equipment", 22), ("Slot", 8), ("Cr", 8), ("Conf", 7), ("When", 14), ("Effect", 66)],
     ["Med-Kit", 1, 20, "C", "Combat", "Cancels the -2 on Stabilize and on treating Bleed & Poison."]),

    ("Chems", "gear",
     [("Chem", 20), ("Cr", 8), ("Conf", 7), ("Duration", 16), ("Effect", 60), ("Dependence", 26)],
     ["Combat Stim", 15, "C", "1 activation", "+1 to melee and ranged attack rolls.",
      "NRV test or gain Dependence"]),

    ("Drones", "gear",
     [("Drone", 20), ("Category", 16), ("Cr", 8), ("Conf", 7), ("Bandwidth", 11),
      ("Profile", 30), ("Effect", 56)],
     ["Recon Spotter", "Recon", 25, "C", 1, "MOV 10\", WND 1", "Spots for the crew within 12\"."]),

    # ---- DEPLOYABLES ---------------------------------------------------
    ("Turrets", "deploy",
     [("Turret", 20), ("Build", 12), ("Cr", 8), ("Conf", 7), ("Range", 9), ("Profile", 60)],
     ["Autoturret", "Complex -1", 10, "B", "18\"", "One shot / round, Injury Damage +3."]),

    ("Mine Chassis", "deploy",
     [("Chassis", 18), ("Build", 12), ("Cr", 8), ("Conf", 7), ("Delivery", 56)],
     ["Pressure", "Standard 0", 5, "C", "Detonates when a model moves within 1\"."]),

    ("Mine Payloads", "deploy",
     [("Payload", 18), ("Cr", 8), ("Conf", 7), ("Effect", 66)],
     ["Explosion", 10, "C", "Injury roll at Damage +3 against every model within 2\"."]),

    ("Traps", "deploy",
     [("Trap", 20), ("Build", 12), ("Cr", 8), ("Conf", 7), ("Trigger & Effect", 66)],
     ["Trip Wire", "Standard 0", 5, "C", "First model to cross is Pinned and stops."]),

    ("Beacons", "deploy",
     [("Beacon", 20), ("Build", 12), ("Cr", 8), ("Conf", 7), ("Aura (within 6\")", 66)],
     ["Munitions Beacon", "Complex -1", 5, "C", "Allies: +1 to the Injury roll."]),

    # ---- COMBAT --------------------------------------------------------
    ("Conditions", "combat",
     [("Condition", 20), ("Group", 20), ("Clock", 22), ("Effect", 66), ("Cleared by", 34)],
     ["Pinned", "Core combat", "Until cleared", "-1 to all rolls; clear with a Move.",
      "Moving, or a Rally"]),

    ("Reactions", "combat",
     [("Reaction", 18), ("Trigger", 40), ("Effect", 62), ("Cost", 18)],
     ["Snap Shot", "Enemy ends a Move or Action in your forward LOS",
      "Make a normal ranged attack at the triggering enemy.", "Ready token"]),

    # ---- TERRAIN -------------------------------------------------------
    ("Terrain Types", "terrain",
     [("Type", 18), ("Movement", 18), ("Cover", 22), ("Tags", 34), ("Notes", 40)],
     ["Ruin", "Difficult", "Light", "Climbable, Searchable", "The default mid-board piece"]),

    ("Terrain Hazards", "terrain",
     [("Hazard", 22), ("Trigger", 34), ("Effect", 56)],
     ["Fire / burning ground", "Entering or starting in it", "Injury roll at Damage +2; gain Fire."]),

    ("Terrain Verbs", "terrain",
     [("Verb", 22), ("Tag", 22), ("Stat", 8), ("Resolution", 18), ("Notes", 40)],
     ["Force door", "Breachable", "STR", "7+", "Attempt is Loud"]),

    ("Infrastructure", "terrain",
     [("Feature", 24), ("Category", 16), ("Verbs", 30), ("Operate", 18),
      ("Effect", 60), ("Damage Keyword", 16)],
     ["Cargo Crane", "Manipulation", "Shift Terrain / Displace", "INT hack or STR manual",
      "Move a Movable piece up to 4\".", "CRUSH"]),

    # ---- CAMPAIGN ------------------------------------------------------
    ("Structures", "campaign",
     [("Structure", 24), ("Category", 14), ("Class", 12), ("Size", 10), ("Power", 8),
      ("Materials", 11), ("Conf", 7), ("What it does", 62)],
     ["Processor", "Convert", "Plant", "3x5", -1, 45, "C", "Materials gatherer — scrap to Materials"]),

    ("Worker Benefits", "campaign",
     [("Structure", 24), ("Benefit", 62), ("Notes", 34)],
     ["Processor", "+1 Materials per Settlement Phase", ""]),

    ("Scenarios", "campaign",
     [("Scenario", 22), ("Shape", 16), ("Objective", 50), ("Scoring", 40), ("Notes", 34)],
     ["Take a Hold", "Control", "Hold the central objective at the end of each round",
      "1 VP per round held", "The most static of the five"]),

    ("Scenario Twists", "campaign",
     [("d6", 6), ("Twist", 40), ("Effect", 62)],
     [1, "Low light", "All ranged attacks beyond 12\" are at -1."]),

    ("Factions", "campaign",
     [("Faction", 22), ("Battlefield Rule", 62), ("Settlement Affinity", 40), ("Notes", 30)],
     ["Military", "Ready token survives a failed Reaction attempt",
      "+1 free Bunkhouse tier at founding", "Needs a number in the rule"]),

    ("Loot Table", "campaign",
     [("d10", 6), ("Result", 66)],
     [2, "+5 Credits"]),

    ("Events", "campaign",
     [("Roll", 8), ("Event", 26), ("When", 16), ("Effect", 62)],
     [3, "Sudden downpour", "Battlefield", "Ranged attacks beyond 18\" are at -1 for one round."]),

    ("Glorious Deeds", "campaign",
     [("Deed", 24), ("Trigger", 56), ("XP", 6), ("Notes", 30)],
     ["Daredevil", "Leap a gap of 3\" or more during an activation", 1, "Once per fighter per battle"]),

    ("Fate Table", "campaign",
     [("Roll", 8), ("Result", 20), ("Effect", 62)],
     [1, "Dead", "The fighter is removed from the roster permanently."]),

    ("Injuries & Scars", "campaign",
     [("Roll", 8), ("Injury", 24), ("Effect", 56), ("Permanent", 12)],
     [4, "Broken arm", "-1 STR for the next battle", "No"]),
]

LEGEND = [
    ("How to use this workbook", ""),
    ("", ""),
    ("One sheet per catalogue.", "Type your entries under the header row."),
    ("Row 2 is a GREY EXAMPLE.", "It shows the expected format. Delete it when you start — it is ignored on import, and the Progress count below includes it until you do."),
    ("Do not rename sheets or headers.", "The importer matches on both. Adding a column at the right-hand end is safe."),
    ("Leave a cell blank if unknown.", "Blank is honest. A guessed number is the defect this rebuild exists to remove."),
    ("", ""),
    ("The Conf column", ""),
    ("A", "Measured, current, statistically significant — trust it for play."),
    ("B", "Measured, but wide CI or single-scenario coverage only — usable, expect movement."),
    ("C", "Derived by rule from an A/B atom, never measured directly — placeholder."),
    ("(blank)", "Not costed yet. This is the one the old catalogues kept hiding."),
    ("", ""),
    ("The Ruled column", ""),
    ("(blank) or Ross", "You decided it. This is the default — leave it blank and it means yours."),
    ("derived", "Worked out from another rule rather than decided. Fine to keep, but it is an inference, not a ruling."),
    ("carried-over", "Brought in from the old system without being re-checked."),
    ("unruled", "Placeholder. Nobody has actually decided this yet."),
    ("Why it exists", "\"Only a Leader ever reaches T3\" was a derived gloss that entered with the master doc on 2026-08-05, restating a July cap table — then got quoted back at you as a design constraint you never set. This column makes that impossible to repeat."),
    ("", ""),
    ("Sheets that are new", ""),
    ("Chems, Drones", "No catalogue exists in the rules today. Both are named in the setting and referenced by built structures."),
    ("Weapon Drawbacks", "Split out from Characteristics — a refund is a price, and they were never tiered."),
    ("Scenario Twists, Glorious Deeds,", "Existed in the rules but never as their own catalogue."),
    ("Fate Table, Injuries & Scars", ""),
]


def build():
    wb = Workbook()
    wb.remove(wb.active)

    thin = Side(style="thin", color="D5DAD8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ---- README --------------------------------------------------------
    rd = wb.create_sheet("README")
    rd.sheet_properties.tabColor = BAND["meta"]
    rd["A1"] = "SETTLEMENTS — CATALOGUE WORKBOOK"
    rd["A1"].font = Font(name="Arial", size=15, bold=True, color=INK)
    rd["A2"] = "Empty tables for every catalogue in the system. Fill them in; nothing here is generated from the current rules."
    rd["A2"].font = Font(name="Arial", size=10, italic=True, color="5A636B")
    r = 4
    for a, b in LEGEND:
        rd.cell(r, 1, a).font = Font(name="Arial", size=10,
                                     bold=not b or a.endswith(":") or a in ("A", "B", "C", "(blank)"))
        rd.cell(r, 2, b).font = Font(name="Arial", size=10)
        rd.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    rd.column_dimensions["A"].width = 34
    rd.column_dimensions["B"].width = 96

    r += 1
    rd.cell(r, 1, "Progress").font = Font(name="Arial", size=12, bold=True, color=INK)
    r += 1
    rd.cell(r, 1, "Catalogue").font = Font(name="Arial", size=10, bold=True)
    rd.cell(r, 2, "Rows").font = Font(name="Arial", size=10, bold=True)
    head_row = r
    r += 1
    first_count = r

    # ---- one sheet per catalogue --------------------------------------
    for name, band, cols, example in SHEETS:
        # Provenance, on every sheet. The failure this exists to stop: a derived
        # observation hardening into a stated tenet and then being quoted back
        # as a ruling. Blank means Ross typed it, so the common case is no work.
        cols = cols + [("Ruled", 11)]
        example = example + ["Ross"]
        ws = wb.create_sheet(name)
        ws.sheet_properties.tabColor = BAND[band]
        fill = PatternFill("solid", fgColor=BAND[band])
        for c, (header, width) in enumerate(cols, 1):
            cell = ws.cell(1, c, header)
            cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
            ws.column_dimensions[get_column_letter(c)].width = width
        for c, val in enumerate(example, 1):
            cell = ws.cell(2, c, val)
            cell.font = Font(name="Arial", size=10, italic=True, color="9AA4AC")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = "A1:%s1" % get_column_letter(len(cols))

        headers = [h for h, _ in cols]
        dvr = DataValidation(type="list", formula1='"Ross,derived,carried-over,unruled"',
                             allow_blank=True, showErrorMessage=False)
        dvr.prompt = ("Blank or Ross = your decision. derived = worked out from another "
                      "rule. carried-over = brought in from the old system, unchecked. "
                      "unruled = placeholder, nobody has decided.")
        dvr.promptTitle = "Who ruled this?"
        ws.add_data_validation(dvr)
        rcol = get_column_letter(headers.index("Ruled") + 1)
        dvr.add("%s2:%s400" % (rcol, rcol))
        if "Conf" in headers:
            dv = DataValidation(type="list", formula1='"A,B,C"', allow_blank=True,
                                showErrorMessage=False)
            dv.prompt = "A measured/significant, B measured but wide, C derived by rule. Blank = not costed."
            dv.promptTitle = "Confidence tier"
            ws.add_data_validation(dv)
            col = get_column_letter(headers.index("Conf") + 1)
            dv.add("%s2:%s400" % (col, col))

        rd.cell(r, 1, name).font = Font(name="Arial", size=10)
        rd.cell(r, 2, "=COUNTA('%s'!A2:A400)" % name).font = Font(name="Arial", size=10)
        r += 1

    rd.cell(r, 1, "TOTAL").font = Font(name="Arial", size=10, bold=True)
    rd.cell(r, 2, "=SUM(B%d:B%d)" % (first_count, r - 1)).font = Font(name="Arial", size=10, bold=True)
    rd.freeze_panes = "A%d" % (head_row + 1)

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    return len(SHEETS)


if __name__ == "__main__":
    n = build()
    print("wrote %s" % OUT)
    print("  %d catalogue sheets + README" % n)
